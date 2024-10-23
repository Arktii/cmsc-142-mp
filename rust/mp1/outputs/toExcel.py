import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def main():
    directory = os.getcwd()
    data = {}

    wb = Workbook()
    ws = wb.active

    create_headers(ws)

    for filename in os.listdir(directory):
        if filename.endswith('.txt'):
            file_path = os.path.join(directory, filename)
            n = int(filename.split('_')[1])

            if n not in data:
                data[n] = []

            with open(file_path, 'r') as f:
                lines = f.readlines()

                solution =  lines[1].strip().split('[')[1].split(']')[0].split(', ')
                solution = ''.join(solution)

                weight = int(lines[2].strip().split(': ')[1])

                value = int(lines[3].strip().split(': ')[1])

                total_time = lines[4].strip().split(': ')[1]
                #theres a weird character when the unit is in microsecs
                total_time = total_time.strip().replace('Â', '')

                if total_time.endswith('µs'):
                    total_time = float(total_time[:-2])
                elif total_time.endswith('ms'):
                    total_time = float(total_time[:-2]) * 1000
                elif total_time.endswith('s'):
                    total_time = float(total_time[:-1]) * 1000000

                data[n].append([total_time, solution, weight, value])

    add_data(ws, data)
    format_center(ws)
    wb.save('output.xlsx')

def add_data(ws, data):
    for n,sets in data.items():
        row = [n]
        for set_data in sets:
            row.extend(set_data)
        ws.append(row)

def create_headers(ws):
    ws.append(['n', 'Set 1', '', '', '', 'Set 2', '', '', '', 'Set 3', '', '', ''])
    ws.append(['', 'Total Time', 'Solution', 'Weight', 'Value', 'Total Time', 'Solution', 'Weight', 'Value', 'Total Time', 'Solution', 'Weight', 'Value'])

    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['F1'].font = Font(bold=True)
    ws['J1'].font = Font(bold=True)

    #A BC FG JK
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 16

    ws.merge_cells('A1:A2')  # Merging for 'n'
    ws.merge_cells('B1:E1')  # Merging for 'Set 1'
    ws.merge_cells('F1:I1')  # Merging for 'Set 2'
    ws.merge_cells('J1:M1')  # Merging for 'Set 3'

def format_center(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


if __name__ == '__main__':
    main()